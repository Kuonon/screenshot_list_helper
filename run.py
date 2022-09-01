import os
import xlrd
import shutil
import smtplib
import datetime
import requests
import openpyxl
import zipfilegbk
from docx import Document
from docx.shared import Cm
from zipfile import ZipFile
from shutil import copyfile
from email.utils import formataddr
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication



info = "\n请选择模式：\n" \
       "\033[32m1.青年大学习(待开发)\033[0m\n" \
       "\033[32m2.每月一学\033[0m"
print(info)
mode = input("您的选择是：")

if mode == "1":
    set_mode = 1
elif mode == "2":
    set_mode = 2
else:
    print("\033[31m模式选择失败，请重新启动\033[0m\n")
    input()

try:
    os.makedirs(r"./raw")
    url_input = input("请输入压缩包链接：")
    url = url_input
    r = requests.get(url)
    with open(r'./raw/latest_raw.zip', 'wb') as code:  # 将压缩包内容写入到 "./raw/" 下，并命名
        code.write(r.content)
        r.close()
        code.close()
    down_success = "\n——————————————————————————————————————————————————————————————————————————————————————————\n" \
                   "\033[32m压缩包下载成功\033[0m\n" \
                   "——————————————————————————————————————————————————————————————————————————————————————————"
    print(down_success)
except:
    down_error = "——————————————————————————————————————————————————————————————————————————————————————————\n"\
                 "\033[31m下载失败，请手动放置压缩包于./raw/目录下，然后按Enter继续\033[0m\n"\
                 "——————————————————————————————————————————————————————————————————————————————————————————"
    print(down_error)
    input()

try:
    zip = r"./raw"
    zip_name = os.listdir(zip)[0]
    read_success = "——————————————————————————————————————————————————————————————————————————————————————————\n"\
                   "检测到压缩文件名称为：\n\033[34m%s\033[0m\n"\
                   "——————————————————————————————————————————————————————————————————————————————————————————" % (zip_name)
    print(read_success)
except:
    read_error = "\n——————————————————————————————————————————————————————————————————————————————————————————\n"\
                 "\033[31m未检测到压缩文件\033[0m\n"\
                 "——————————————————————————————————————————————————————————————————————————————————————————"
    print(read_error)
    input()

try:
    os.makedirs(r"./screenshot")
    zip_name_ext = "./raw/" + zip_name
    dst_dir = "./screenshot"
    fz = zipfilegbk.ZipFile(zip_name_ext, 'r')
    for file in fz.namelist():
        fz.extract(file, dst_dir)
    fz.close()
    ext_success = "——————————————————————————————————————————————————————————————————————————————————————————\n"\
                   "\033[32m解压成功\033[0m\n"\

    print(ext_success)
    input("\033[33m若有补交截图，请命名后放置于./screenshot/目录下，按Enter继续。\033[0m\n"
          "——————————————————————————————————————————————————————————————————————————————————————————")
except:
    ext_error = "——————————————————————————————————————————————————————————————————————————————————————————\n"\
                 "\033[31m解压失败\033[0m\n"\
                 "——————————————————————————————————————————————————————————————————————————————————————————"
    print(ext_error)
    input()

try:
    jpg = r"./screenshot" #图片路径
    jpg_name_raw = os.listdir(jpg) #所有图片文件名
    file_count = len(jpg_name_raw) #获取文件个数

    i = 0
    name_list = []  # 创建读取名字列表
    name_attach = []  # 创建读取名字后缀列表
    while i != len(jpg_name_raw):
        jpg_name_list = jpg_name_raw[i] #提取第i个文件名
        jpg_name = os.path.splitext(jpg_name_list)[0] #提取第i个文件的文件名
        jpg_attach = os.path.splitext(jpg_name_list)[1] #提取第i个文件的后缀名
        name_list.append(jpg_name) #写入名字列表
        name_attach.append(jpg_attach)  # 写入后缀列表
        i = i + 1
    pass
    jpg_count_success = "——————————————————————————————————————————————————————————————————————————————————————————\n"\
                        "\033[32m截图计数成功\033[0m\n"\
                        "——————————————————————————————————————————————————————————————————————————————————————————"
    print(jpg_count_success)
except:
    jpg_count_error = "——————————————————————————————————————————————————————————————————————————————————————————\n" \
                      "\033[31m截图计数失败\033[0m\n" \
                      "——————————————————————————————————————————————————————————————————————————————————————————"
    print(jpg_count_error)
    input()

try:
    wb = xlrd.open_workbook('./list.xls') #读取已知名单
    sh = wb.sheet_by_name('list') #读取表单
    list_all = sh.col_values(1) #读取姓名列

    list_unupload = set(list_all).difference(set(name_list)) #未交截图名单
    list_upload = set(list_all).intersection((set(name_list))) #已交截图名单
    list_error = set(name_list).difference(set(list_all)) #姓名错误名单

    count_unupload = len(list_unupload) #未交截图名单计数
    count_upload = len(list_upload) #已交截图名单计数
    count_error = len(list_error) #姓名错误名单计数

    if count_error == 0:
        list_error = "无"
    else:
        pass

    list_count_success = "——————————————————————————————————————————————————————————————————————————————————————————\n"\
                         "名单统计：\n" \
                         "共收集到(\033[32m%s\033[0m/57)份截图，核验后有效截图\033[32m%s\033[0m张\n" \
                         "其中未交截图的同学有\033[31m%s\033[0m人，分别是\033[34m%s\033[0m同学\n" \
                         "其中姓名错误者有\033[33m%s\033[0m人，错误值：\033[33m%s\033[0m\n"\
                         "——————————————————————————————————————————————————————————————————————————————————————————" % (file_count, count_upload, count_unupload, list_unupload, count_error, list_error)
    print(list_count_success)
except:
    list_count_error = "——————————————————————————————————————————————————————————————————————————————————————————\n" \
                      "\033[31m名单统计失败\033[0m\n" \
                      "——————————————————————————————————————————————————————————————————————————————————————————"
    print(list_count_error)
    input()

try:
    a = 0
    list_upload_list = list(list_upload)
    name_list_rank = []  # 顺序实际姓名
    id_list_rank = []  # 顺序实际学号
    while a != 57:
        list_name_singular = sh.cell_value(rowx=a, colx=1) #提取第a个总名单里的名字
        list_id_singular = sh.cell_value(rowx=a, colx=2) #提取第a个总名单里的学号
        list_name_singular_list = [] #单名字列表
        list_name_singular_list.append(list_name_singular) #添加单个名字进入单名字列表
        search_condition = set(list_upload_list).intersection((set(list_name_singular_list))) #与实际名单取交集，检验是否存在
        search_condition_singular = ''.join(search_condition) #单个集合转为字符串
        search_condition_count = len(search_condition) #检验交集长度，结果为0或1
        if search_condition_count == 1: #名字存在，交了截图
            name_list_rank.append(search_condition_singular) #那么添加名字入列
            id_list_rank.append(list_id_singular) #那么添加学号入列
        else: #名字不存在，没交截图
            pass
        a = a + 1
    list_drop_success = "——————————————————————————————————————————————————————————————————————————————————————————\n" \
                        "\033[32m已交截图的(姓名-学号)字段提取成功\033[0m\n" \
                        "——————————————————————————————————————————————————————————————————————————————————————————"
    print(list_drop_success)
except:
    list_drop_error = "——————————————————————————————————————————————————————————————————————————————————————————\n" \
                       "\033[31m已交截图的(姓名-学号)字段提取失败\033[0m\n" \
                       "——————————————————————————————————————————————————————————————————————————————————————————"
    print(list_drop_error)
    input()

try:
    file_name = "20影本1班（" + str(count_upload) + "份）" #生成目录文件夹名字
    os.makedirs(r"./" + file_name) #生成目录文件夹路径
    copyfile("./formwork/monthly_study/list.xlsx", "./" + file_name + "/" + "20影本1班+每月一学截图名单.xlsx") #复制xlsx表格
    copyfile("./formwork/monthly_study/img.docx", "./" + file_name + "/" + "20影本1班+每月一学截图（" + str(count_upload) + "份）.docx") #复制docx表格
    catalog_build_success = "——————————————————————————————————————————————————————————————————————————————————————————\n" \
                        "\033[32m提交文件夹和提交模板创建成功\033[0m\n" \
                        "——————————————————————————————————————————————————————————————————————————————————————————"
    print(catalog_build_success)
except:
    catalog_build_error = "——————————————————————————————————————————————————————————————————————————————————————————\n" \
                      "\033[31m提交文件夹和提交模板创建失败\033[0m\n" \
                      "——————————————————————————————————————————————————————————————————————————————————————————"
    print(catalog_build_error)
    input()

try:
    workbook = openpyxl.load_workbook("./" + file_name + "/" + "20影本1班+每月一学截图名单.xlsx") #打开xlsx表格
    sheet = workbook['Sheet1'] #找到Sheet1表单
    b = 0
    while b != count_upload:
        sheet.cell(row=2+b, column=1, value=1+b) #序号列
        sheet.cell(row=2+b, column=2, value=name_list_rank[b]) #姓名列
        sheet.cell(row=2+b, column=3, value='20影本1班') #班级列
        sheet.cell(row=2+b, column=4, value=id_list_rank[b]) #学号列

        b = b + 1

    workbook.save("./" + file_name + "/" + "20影本1班+每月一学截图名单.xlsx")
    xlsx_write_success = "——————————————————————————————————————————————————————————————————————————————————————————\n" \
                        "\033[32m生成excel名单文件成功\033[0m\n" \
                        "——————————————————————————————————————————————————————————————————————————————————————————"
    print(xlsx_write_success)
except:
    xlsx_write_error = "——————————————————————————————————————————————————————————————————————————————————————————\n" \
                      "\033[31m生成excel名单文件失败\033[0m\n" \
                      "——————————————————————————————————————————————————————————————————————————————————————————"
    print(xlsx_write_error)
    input()

try:
    c = 0
    document = Document("./" + file_name + "/" + "20影本1班+每月一学截图（" + str(count_upload) + "份）.docx") #打开docx文件
    while c != count_upload:
        link_attach = name_list.index(name_list_rank[c]) #获取截图列表中某个名字的索引
        name_attach_rank = name_attach[link_attach] #通过索引得到某个名字的后缀
        document.add_picture("./screenshot/" + name_list_rank[c] + name_attach_rank, width=Cm(10)) #顺序名字 + 顺序后缀 + 宽度10厘米
        if c == count_upload - 1: #若截图到最后一张时
            pass #不加分页符
        else:
            document.add_page_break() #加分页符
        c = c + 1

    document.save("./" + file_name + "/" + "20影本1班+每月一学截图（" + str(count_upload) + "份）.docx")

    docx_write_success = "——————————————————————————————————————————————————————————————————————————————————————————\n" \
                        "\033[32m生成word截图名单文件成功\033[0m\n" \
                        "——————————————————————————————————————————————————————————————————————————————————————————"
    print(docx_write_success)
except:
    docx_write_error = "——————————————————————————————————————————————————————————————————————————————————————————\n" \
                      "\033[31m生成word截图名单文件失败\033[0m\n" \
                      "——————————————————————————————————————————————————————————————————————————————————————————"
    print(docx_write_error)
    input()

try:
    name_pickup = "20影本1班（" + str(count_upload) + "份）.zip"
    with ZipFile(name_pickup, 'w') as handle: #新建截图打包文件
        os.chdir("./") # 首先切到根目录
        handle.write("20影本1班（" + str(count_upload) + "份）/20影本1班+每月一学截图名单.xlsx") #写入xlsx表单
        handle.write("20影本1班（" + str(count_upload) + "份）/20影本1班+每月一学截图（" + str(count_upload) + "份）.docx") #写入docx表单
    pickup_success = "——————————————————————————————————————————————————————————————————————————————————————————\n" \
                        "\033[32m文件夹打包压缩成功\033[0m\n" \
                        "——————————————————————————————————————————————————————————————————————————————————————————"
    print(pickup_success)
except:
    pickup_error = "——————————————————————————————————————————————————————————————————————————————————————————\n" \
                      "\033[31m文件夹打包压缩失败\033[0m\n" \
                      "——————————————————————————————————————————————————————————————————————————————————————————"
    print(pickup_error)
    input()

confirm = "——————————————————————————————————————————————————————————————————————————————————————————\n" \
                        "\033[32m截图处理完毕，请手动确认无误后输入验证码\033[0m\033[33m1390\033[0m\033[32m确认发送\033[0m\n" \
                        "——————————————————————————————————————————————————————————————————————————————————————————"
print(confirm)
confirm_code = input("确认码：")
if confirm_code == "1390":
    try:
        # 设置邮箱的域名
        HOST = 'smtp.qq.com'
        # 设置邮件标题
        month = datetime.date.today().month
        SUBJECT = "20影本1班-每月一学(" + str(month) + "月)截图+名单（" + str(count_upload) + "份）"
        # 设置发件人邮箱
        FROM = '2858493933@qq.com'
        # 设置收件人邮箱，可以同时发送到多个邮箱（用list）
        TO = '2858493933@qq.com'
        # 设置附件模式
        message = MIMEMultipart('mixed')

        ''' 添加正文 '''
        content = ""
        content_msg = MIMEText(content)
        message.attach(content_msg)
        message['Subject'] = SUBJECT# 邮件的主题，也可以说是标题
        message['From'] = formataddr(["20影本1班宣传委员", FROM]) # 括号里的对应发件人邮箱昵称、发件人邮箱账号
        message['To'] = formataddr(["user", TO])
        zip = "20影本1班（" + str(count_upload) + "份）.zip"
        zip_file = MIMEApplication(open(zip, 'rb').read())
        zip_file.add_header('Content-Disposition', 'attachment', filename=zip)
        message.attach(zip_file)

        # 获取SSL证书
        email_client = smtplib.SMTP_SSL(host='smtp.qq.com')
        # 设置域名和端口，端口为465
        email_client.connect(HOST, "465")
        # 邮箱授权码
        email_client.login(FROM, 'oynnhfaewdekddij')
        email_client.sendmail(from_addr=FROM, to_addrs=TO.split(','), msg=message.as_string())
        # 关闭邮件发送客户端
        email_client.quit()
        send_success = "——————————————————————————————————————————————————————————————————————————————————————————\n" \
                            "\033[32m邮件发送成功\033[0m\n" \
                            "——————————————————————————————————————————————————————————————————————————————————————————"
        print(send_success)
    except:
        send_error = "——————————————————————————————————————————————————————————————————————————————————————————\n" \
                          "\033[31m邮件发送失败\033[0m\n" \
                          "——————————————————————————————————————————————————————————————————————————————————————————"
        print(send_error)
        input()

else:
    confirm_error = "——————————————————————————————————————————————————————————————————————————————————————————\n" \
              "\033[31m验证码错误!\033[0m\n" \
              "——————————————————————————————————————————————————————————————————————————————————————————"
    print(confirm_error)

try:
    #datetime.date.today().month
    file_name_finish = str(datetime.date.today()) + " " + str(9) + "月每月一学" #生成目录文件夹名字
    os.makedirs(r"C:/Users/Kuonon/OneDrive/桌面/宣委工作/" + file_name_finish) #生成目录文件夹路径
    shutil.move("./20影本1班（" + str(count_upload) + "份）", "C:/Users/Kuonon/OneDrive/桌面/宣委工作/" + file_name_finish)
    shutil.move("./20影本1班（" + str(count_upload) + "份）.zip", "C:/Users/Kuonon/OneDrive/桌面/宣委工作/" + file_name_finish)
    shutil.rmtree("./screenshot")
    shutil.rmtree("./raw")
    delete_success = "——————————————————————————————————————————————————————————————————————————————————————————\n" \
                        "\033[32m残余文件清理成功\033[0m\n" \
                        "——————————————————————————————————————————————————————————————————————————————————————————"
    print(delete_success)
except:
    delete_error = "——————————————————————————————————————————————————————————————————————————————————————————\n" \
                      "\033[31m残余文件清理失败,请手动删除\033[0m\n" \
                      "——————————————————————————————————————————————————————————————————————————————————————————"
    print(delete_error)
    input()
